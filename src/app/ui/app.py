import os
import asyncio
import psutil
import threading
import tkinter as tk
from tkinter import filedialog
import openpyxl
import requests
from imgui_bundle import imgui, immapp, hello_imgui, icons_fontawesome
from app.core.browser_manager import browser_manager
from app.core.queue_manager import queue_manager
from app.utils.logger import logger
from app.utils.config_manager import config_manager

class WhatsAppUI:
    def __init__(self, loop):
        self.loop = loop
        self.new_client_name = ""
        self.show_add_modal = False
        self.show_config_client = None
        self.show_delete_confirm = None
        self.active_tab = "LOGS"
        self.queue_counts = {}
        self.all_redis_queues = []
        self.sessions = config_manager.get_client_list()
        self.ram_usage = 0
        self.cpu_usage = 0
        self.process = psutil.Process(os.getpid())
        
        # Ycloud Variables
        self.ycloud_contacts = []
        self.ycloud_sent_count = 0
        self.ycloud_failed_count = 0
        self.ycloud_total_cost_usd = 0.0
        self.ycloud_is_sending = False
        self.show_ycloud_config = False
        
        # Load configs
        self.ycloud_api_key = config_manager.get_global("ycloud_api_key", "")
        self.ycloud_from = config_manager.get_global("ycloud_from", "+51963828458")
        self.ycloud_url = config_manager.get_global("ycloud_url", "https://api.ycloud.com/v2/whatsapp/messages/sendDirectly")
        self.ycloud_cost_usd = config_manager.get_global("ycloud_cost_usd", 0.02)
        self.ycloud_exchange_pen = config_manager.get_global("ycloud_exchange_pen", 3.46)
        self.ycloud_colegio = config_manager.get_global("ycloud_colegio", "Mi Colegio")
        self.ycloud_numero = config_manager.get_global("ycloud_numero", "999888777")
        self.ycloud_template = config_manager.get_global("ycloud_template", "🚨🇨🇴🇱🇪✅ \nEstimados padres de familia 👨‍👩‍👧‍👦:\nEste año, *{Colegio}* viene implementando un sistema de control de *INGRESO* y *SALIDA* de estudiantes mediante credenciales escolares.\n\n📌 Guarde el siguiente número como:\n👉 colecheck – *{Numero}*  (escribir a ese whatsapp)\n\n📩 Luego envíe:\n\n🏫 Colegio: *{Colegio}*\n🎓 Grado y sección: *5-A* \n👦 Estudiante: *Yhon Yucra Castro*\n\n🎫 *Verifique que su hijo(a) lleve siempre su credencial, ya que las notificaciones dependen de su uso al ingresar y salir del colegio.*\n\n💬 Para mantener activo el servicio, responda o reaccione 👍 a los mensajes enviados.")


    async def update_data(self):
        while True:
            await queue_manager.connect()
            for name in self.sessions:
                self.queue_counts[name] = await queue_manager.get_queue_size(name)
            self.all_redis_queues = await queue_manager.get_all_queues()
            try:
                mem = self.process.memory_full_info().uss / (1024 * 1024)
                for child in self.process.children(recursive=True):
                    mem += child.memory_full_info().uss / (1024 * 1024)
                self.ram_usage = mem
                self.cpu_usage = psutil.cpu_percent(interval=None)
            except: pass
            await asyncio.sleep(1)

    def load_excel(self):
        def _load():
            root = tk.Tk()
            root.withdraw()
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            root.destroy()
            if not file_path:
                return
            
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # Buscar encabezados
                headers = {}
                for col_idx, cell in enumerate(ws[1], start=1):
                    if cell.value:
                        headers[cell.value.strip().lower()] = col_idx
                
                new_contacts = []
                for row_idx in range(2, ws.max_row + 1):
                    def get_val(key):
                        for h in headers:
                            if key in h:
                                return str(ws.cell(row=row_idx, column=headers[h]).value or "")
                        return ""
                    
                    whatsapp = get_val("whatsapp 1") or get_val("whassap 1") or get_val("whatsapp")
                    if whatsapp and whatsapp != "None":
                        whatsapp = "".join(filter(str.isdigit, whatsapp))
                        if not whatsapp.startswith("51"):
                            whatsapp = "51" + whatsapp
                        whatsapp = "+" + whatsapp
                        
                        estudiante = get_val("estudiante") or get_val("nombre")
                        grado = get_val("grado")
                        seccion = get_val("secci")
                        nro = get_val("n°") or get_val("nro") or get_val("no")
                        
                        new_contacts.append({
                            "nro": nro,
                            "estudiante": estudiante,
                            "grado": grado,
                            "seccion": seccion,
                            "whatsapp": whatsapp,
                            "status": "pending"
                        })
                self.ycloud_contacts = new_contacts
                self.ycloud_sent_count = 0
                self.ycloud_failed_count = 0
            except Exception as e:
                print(f"Error loading Excel: {e}")
                
        threading.Thread(target=_load, daemon=True).start()

    def send_ycloud_messages(self):
        if self.ycloud_is_sending: return
        self.ycloud_is_sending = True
        self.ycloud_sent_count = 0
        self.ycloud_failed_count = 0
        self.ycloud_total_cost_usd = 0.0
        
        config_manager.settings["global"]["ycloud_api_key"] = self.ycloud_api_key
        config_manager.settings["global"]["ycloud_from"] = self.ycloud_from
        config_manager.settings["global"]["ycloud_url"] = self.ycloud_url
        config_manager.settings["global"]["ycloud_cost_usd"] = self.ycloud_cost_usd
        config_manager.settings["global"]["ycloud_exchange_pen"] = self.ycloud_exchange_pen
        config_manager.settings["global"]["ycloud_colegio"] = self.ycloud_colegio
        config_manager.settings["global"]["ycloud_numero"] = self.ycloud_numero
        config_manager.settings["global"]["ycloud_template"] = self.ycloud_template
        config_manager.save()
        
        def _send():
            headers = {
                "X-API-Key": self.ycloud_api_key,
                "accept": "application/json",
                "content-type": "application/json"
            }
            
            for contact in self.ycloud_contacts:
                if contact["status"] == "sent": continue
                
                body_text = self.ycloud_template.replace("{Estudiante}", contact["estudiante"])
                body_text = body_text.replace("{Grado}", contact["grado"])
                body_text = body_text.replace("{Sección}", contact["seccion"])
                body_text = body_text.replace("{Colegio}", self.ycloud_colegio)
                body_text = body_text.replace("{Numero}", self.ycloud_numero)
                
                payload = {
                    "type": "text",
                    "text": {
                        "body": body_text,
                        "preview_url": False
                    },
                    "from": self.ycloud_from,
                    "to": contact["whatsapp"]
                }
                
                try:
                    res = requests.post(self.ycloud_url, json=payload, headers=headers, timeout=10)
                    if res.status_code in [200, 202]:
                        contact["status"] = "sent"
                        self.ycloud_sent_count += 1
                        self.ycloud_total_cost_usd += self.ycloud_cost_usd
                    else:
                        contact["status"] = "failed"
                        self.ycloud_failed_count += 1
                        print(f"Failed to send to {contact['whatsapp']}: {res.text}")
                except Exception as e:
                    contact["status"] = "failed"
                    self.ycloud_failed_count += 1
                    print(f"Error sending to {contact['whatsapp']}: {e}")
                    
            self.ycloud_is_sending = False
            
        threading.Thread(target=_send, daemon=True).start()

    def draw(self):
        if not hasattr(self, 'update_task_started'):
            asyncio.run_coroutine_threadsafe(self.update_data(), self.loop)
            self.update_task_started = True

        imgui.push_style_color(imgui.Col_.window_bg, (0.07, 0.08, 0.1, 1.0))
        imgui.push_style_color(imgui.Col_.child_bg, (0.09, 0.1, 0.12, 1.0))
        imgui.push_style_color(imgui.Col_.text, (0.85, 0.88, 0.9, 1.0))
        imgui.push_style_color(imgui.Col_.button, (0.15, 0.17, 0.22, 1.0))

        io = imgui.get_io()
        imgui.set_next_window_pos((0, 0))
        imgui.set_next_window_size(io.display_size)
        imgui.begin("Main", None, imgui.WindowFlags_.no_title_bar | imgui.WindowFlags_.no_resize)

        # --- SIDEBAR ---
        # Hacer el sidebar responsivo: 28% de la ventana actual, mínimo 280px
        window_width = imgui.get_window_width()
        sidebar_width = max(280, int(window_width * 0.28))
        imgui.begin_child("Sidebar", (sidebar_width, -60), True)
        if imgui.button(f"{icons_fontawesome.ICON_FA_ROCKET}  START ALL", (sidebar_width - 20, 35)):
            for name in self.sessions:
                cfg = config_manager.get_client_config(name)
                if cfg.get("enabled", True):
                    asyncio.run_coroutine_threadsafe(browser_manager.get_page(name), self.loop)
                    asyncio.run_coroutine_threadsafe(queue_manager.start_worker(name), self.loop)
        
        imgui.spacing(); imgui.separator(); imgui.spacing()
        for name in self.sessions:
            imgui.push_id(name)
            
            # Checkbox para habilitar/deshabilitar
            cfg = config_manager.get_client_config(name)
            enabled = cfg.get("enabled", True)
            
            status = browser_manager.get_status(name)
            
            status_icon = icons_fontawesome.ICON_FA_QUESTION_CIRCLE
            status_color = (0.5, 0.5, 0.5, 1.0)
            
            if not enabled:
                status_icon = icons_fontawesome.ICON_FA_POWER_OFF
                status_color = (0.3, 0.3, 0.3, 1.0)
            elif status == "READY":
                status_icon = icons_fontawesome.ICON_FA_CHECK_CIRCLE
                status_color = (0.1, 0.8, 0.4, 1.0)
            elif status == "STARTING":
                status_icon = icons_fontawesome.ICON_FA_CIRCLE_NOTCH
                status_color = (1.0, 0.6, 0.0, 1.0)
            elif status == "ERROR":
                status_icon = icons_fontawesome.ICON_FA_EXCLAMATION_TRIANGLE
                status_color = (1.0, 0.2, 0.2, 1.0)

            imgui.begin_group()
            
            # Línea 1: Icono de estado + Nombre + Contadores
            imgui.text_colored(status_color, status_icon)
            imgui.same_line()
            
            if not enabled: imgui.push_style_color(imgui.Col_.text, (0.5, 0.5, 0.5, 1.0))
            imgui.text(name)
            if not enabled: imgui.pop_style_color()
            
            count = self.queue_counts.get(name, 0)
            sent = logger.get_account_stats(name)
            
            imgui.same_line(sidebar_width - 85)
            imgui.text_colored((0.2, 0.9, 0.5, 1), str(sent)) # Enviados
            imgui.same_line(sidebar_width - 40)
            imgui.text_colored((1, 0.8, 0.2, 1), str(count)) # En cola
            
            # Línea 2: Toggle "On" (Activo) y Botones de acción
            imgui.set_cursor_pos_x(10)
            c_en, enabled = imgui.checkbox("On", enabled)
            if c_en:
                config_manager.set_client_config(name, {"enabled": enabled})
            
            imgui.same_line()
            if imgui.small_button(f"{icons_fontawesome.ICON_FA_QRCODE} AUTH"):
                asyncio.run_coroutine_threadsafe(browser_manager.get_page(name, visible=True), self.loop)
            
            imgui.same_line()
            is_paused = name in queue_manager.paused_workers
            icon_play_pause = icons_fontawesome.ICON_FA_PLAY if is_paused else icons_fontawesome.ICON_FA_PAUSE
            if imgui.small_button(icon_play_pause):
                queue_manager.toggle_pause(name)
                if is_paused:
                    asyncio.run_coroutine_threadsafe(queue_manager.start_worker(name), self.loop)
            
            imgui.same_line()
            if imgui.small_button(icons_fontawesome.ICON_FA_COG):
                self.show_config_client = name
            
            imgui.same_line()
            imgui.push_style_color(imgui.Col_.text, (1, 0.3, 0.3, 1))
            if imgui.small_button(icons_fontawesome.ICON_FA_TRASH):
                self.show_delete_confirm = name
            imgui.pop_style_color()
            
            imgui.end_group()
            
            imgui.pop_id(); imgui.spacing(); imgui.separator()
        
        imgui.set_cursor_pos((10, imgui.get_window_height() - 55))
        if imgui.button(f"{icons_fontawesome.ICON_FA_PLUS_CIRCLE}  New Client", (sidebar_width - 20, 30)):
            self.show_add_modal = True
        imgui.end_child()

        imgui.same_line()

        # --- CONTENIDO ---
        imgui.begin_child("Content", (0, -60), False)
        if imgui.button(f"{icons_fontawesome.ICON_FA_LIST_ALT}  LOGS"):
            self.active_tab = "LOGS"
        imgui.same_line()
        if imgui.button(f"{icons_fontawesome.ICON_FA_COGS}  GLOBAL CONFIG"):
            self.active_tab = "CONFIG"
        imgui.same_line()
        if imgui.button(f"{icons_fontawesome.ICON_FA_PAPER_PLANE}  YCLOUD"):
            self.active_tab = "YCLOUD"
        imgui.separator()

        if self.active_tab == "LOGS":
            # --- LOGS HEADER STATS ---
            stats = logger.get_stats()
            total = stats["morning"] + stats["afternoon"]
            
            imgui.begin_child("LogStats", (0, 40), True)
            imgui.text(f"{icons_fontawesome.ICON_FA_CHART_BAR}  HOY:")
            imgui.same_line(); imgui.text_colored((0.4, 0.8, 1.0, 1.0), f"Mañana: {stats['morning']}")
            imgui.same_line(); imgui.text(" | ")
            imgui.same_line(); imgui.text_colored((1.0, 0.7, 0.3, 1.0), f"Tarde: {stats['afternoon']}")
            imgui.same_line(); imgui.text(" | ")
            imgui.same_line(); imgui.text_colored((0.2, 0.9, 0.5, 1.0), f"Total: {total}")
            imgui.end_child()
            
            imgui.spacing()
            
            imgui.begin_child("LogsInner", (0, 0), False)
            
            # Table-like structure for logs
            if imgui.begin_table("LogTable", 3, imgui.TableFlags_.resizable | imgui.TableFlags_.scroll_y):
                imgui.table_setup_column("Time", imgui.TableColumnFlags_.width_fixed, 70)
                imgui.table_setup_column("Account", imgui.TableColumnFlags_.width_fixed, 100)
                imgui.table_setup_column("Message", imgui.TableColumnFlags_.width_stretch)
                
                for log in logger.get_logs():
                    imgui.table_next_row()
                    
                    # Col 1: Time
                    imgui.table_next_column()
                    imgui.text_disabled(log["time"])
                    
                    # Col 2: Account
                    imgui.table_next_column()
                    if log["account"]:
                        # Dot indicator
                        draw_list = imgui.get_window_draw_list()
                        pos = imgui.get_cursor_screen_pos()
                        draw_list.add_circle_filled((pos.x + 5, pos.y + 10), 3, imgui.get_color_u32(log["account_color"]))
                        
                        imgui.set_cursor_pos_x(imgui.get_cursor_pos_x() + 15)
                        imgui.text_colored(log["account_color"], log["account"])
                    else:
                        imgui.text_disabled("system")
                        
                    # Col 3: Message
                    imgui.table_next_column()
                    level = log["level"]
                    msg = log["message"]
                    
                    # Custom icon coloring/rendering
                    if level == "ERROR": 
                        imgui.text_colored((1, 0.4, 0.4, 1), f"{icons_fontawesome.ICON_FA_EXCLAMATION_TRIANGLE} {msg}")
                    elif level == "SUCCESS": 
                        imgui.text_colored((0.2, 0.9, 0.5, 1), f"{icons_fontawesome.ICON_FA_CHECK_CIRCLE} {msg}")
                    elif level == "WARN": 
                        imgui.text_colored((1, 0.8, 0.2, 1), f"{icons_fontawesome.ICON_FA_EXCLAMATION_CIRCLE} {msg}")
                    elif level == "DEBUG": 
                        imgui.text_disabled(f"{icons_fontawesome.ICON_FA_SEARCH} {msg}")
                    else: 
                        imgui.text(f"{icons_fontawesome.ICON_FA_INFO_CIRCLE} {msg}")
                
                if imgui.get_scroll_y() >= imgui.get_scroll_max_y():
                    imgui.set_scroll_here_y(1.0)
                
                imgui.end_table()
            
            imgui.end_child()
        
        elif self.active_tab == "CONFIG":
            imgui.text_colored((0.3, 0.7, 1.0, 1.0), "1. PAUSE & DELAY SETTINGS")
            
            # Leer el estado actual de la memoria
            min_d = config_manager.get_global("min_delay", 2)
            max_d = config_manager.get_global("max_delay", 5)
            b_size = config_manager.get_global("batch_size", 20)
            b_pause = config_manager.get_global("batch_pause", 60)
            
            # Capturar los cambios en tiempo real, permitiendo bajar hasta 0
            c1, min_d = imgui.slider_int("Min Delay", min_d, 0, 10)
            c2, max_d = imgui.slider_int("Max Delay", max_d, 0, 60)
            c3, b_size = imgui.slider_int("Batch Size", b_size, 1, 500)
            c4, b_pause = imgui.slider_int("Batch Pause", b_pause, 0, 600)
            
            # Si el usuario mueve un slider, actualizar la memoria instantáneamente
            # Esto evita que el slider vuelva a saltar a su posición original por el bucle de renderizado
            if c1 or c2 or c3 or c4:
                config_manager.settings["global"].update({
                    "min_delay": min_d, "max_delay": max_d, 
                    "batch_size": b_size, "batch_pause": b_pause
                })
            
            if imgui.button("SAVE CONFIG"):
                config_manager.save() # Guarda en el JSON permanentemente
                logger.success("Configuración guardada en config.json")
            
            imgui.spacing(); imgui.separator(); imgui.spacing()
            imgui.text_colored((0.3, 0.7, 1.0, 1.0), f"{icons_fontawesome.ICON_FA_SYNC}  2. QUEUE ORCHESTRATOR")
            if imgui.button(f"{icons_fontawesome.ICON_FA_SYNC_ALT}  SYNC & RESUME ALL REDIS QUEUES", (350, 40)):
                for qn, _ in self.all_redis_queues:
                    if qn in queue_manager.paused_workers:
                        queue_manager.toggle_pause(qn)
                    asyncio.run_coroutine_threadsafe(queue_manager.start_worker(qn), self.loop)
            
            imgui.spacing()
            imgui.text("Current Redis Queues:")
            imgui.columns(4, "qtable")
            imgui.text("Name"); imgui.next_column()
            imgui.text("Queue"); imgui.next_column()
            imgui.text("Sent Today"); imgui.next_column()
            imgui.text("Action"); imgui.next_column()
            imgui.separator()
            for qn, qs in self.all_redis_queues:
                imgui.text(qn); imgui.next_column()
                imgui.text_colored((1, 0.8, 0.2, 1), str(qs)); imgui.next_column()
                
                # Sent count per client
                sent_count = logger.get_account_stats(qn)
                imgui.text_colored((0.2, 0.9, 0.5, 1), str(sent_count)); imgui.next_column()
                
                is_p = qn in queue_manager.paused_workers
                if imgui.small_button(f"{'RESUME' if is_p else 'PAUSE'}##{qn}"):
                    queue_manager.toggle_pause(qn)
                imgui.next_column()
            imgui.columns(1)
            
            imgui.spacing(); imgui.separator(); imgui.spacing()
            imgui.text_colored((0.3, 0.7, 1.0, 1.0), "3. WELCOME MESSAGE OVERRIDE")
            
            override_val = config_manager.get_global("override_welcome", False)
            custom_msg = config_manager.get_global("custom_welcome_msg", "Escribe tu mensaje aquí...")
            
            c5, override_val = imgui.checkbox("Sobrescribir Mensaje de Bienvenida", override_val)
            
            if override_val:
                imgui.text_disabled("Usa las variables: {usuario}, {contrasena}, {url}")
                c6, custom_msg = imgui.input_text_multiline("##custommsg", custom_msg, (imgui.get_window_width() - 30, 80))
            else:
                c6 = False

            if c5 or c6:
                config_manager.settings["global"]["override_welcome"] = override_val
                config_manager.settings["global"]["custom_welcome_msg"] = custom_msg

        elif self.active_tab == "YCLOUD":
            if imgui.button(f"{icons_fontawesome.ICON_FA_COG}  YCLOUD CONFIGURATION"):
                self.show_ycloud_config = True
                
            imgui.spacing(); imgui.separator(); imgui.spacing()
            imgui.text_colored((0.3, 0.7, 1.0, 1.0), "YCLOUD EXCEL DATA & SENDING")
            
            if imgui.button(f"{icons_fontawesome.ICON_FA_FILE_EXCEL}  LOAD EXCEL"):
                self.load_excel()
            imgui.same_line()
            
            if self.ycloud_is_sending:
                imgui.text_colored((1.0, 0.6, 0.0, 1.0), "Sending in progress...")
            else:
                if imgui.button(f"{icons_fontawesome.ICON_FA_PAPER_PLANE}  START SENDING"):
                    self.send_ycloud_messages()
            
            total_loaded = len(self.ycloud_contacts)
            cost_pen = self.ycloud_total_cost_usd * self.ycloud_exchange_pen
            imgui.text(f"Loaded: {total_loaded} | ")
            imgui.same_line(); imgui.text_colored((0.2, 0.9, 0.5, 1.0), f"Sent: {self.ycloud_sent_count} | ")
            imgui.same_line(); imgui.text_colored((1.0, 0.2, 0.2, 1.0), f"Failed: {self.ycloud_failed_count} | ")
            imgui.same_line(); imgui.text_colored((1.0, 0.8, 0.2, 1.0), f"Cost: ${self.ycloud_total_cost_usd:.2f} USD (S/ {cost_pen:.2f} PEN)")
            
            imgui.spacing()
            imgui.begin_child("YcloudTable", (0, 0), True)
            if imgui.begin_table("YCTable", 6, imgui.TableFlags_.resizable | imgui.TableFlags_.scroll_y | imgui.TableFlags_.borders):
                imgui.table_setup_column("N°", imgui.TableColumnFlags_.width_fixed, 30)
                imgui.table_setup_column("Estudiante")
                imgui.table_setup_column("Grado")
                imgui.table_setup_column("Sección")
                imgui.table_setup_column("Whatsapp")
                imgui.table_setup_column("Status", imgui.TableColumnFlags_.width_fixed, 60)
                imgui.table_headers_row()
                
                for contact in self.ycloud_contacts:
                    imgui.table_next_row()
                    imgui.table_next_column(); imgui.text(str(contact["nro"]))
                    imgui.table_next_column(); imgui.text(str(contact["estudiante"]))
                    imgui.table_next_column(); imgui.text(str(contact["grado"]))
                    imgui.table_next_column(); imgui.text(str(contact["seccion"]))
                    imgui.table_next_column(); imgui.text(str(contact["whatsapp"]))
                    imgui.table_next_column()
                    status = contact["status"]
                    if status == "pending":
                        imgui.text_colored((0.5, 0.5, 0.5, 1.0), "Pending")
                    elif status == "sent":
                        imgui.text_colored((0.2, 0.9, 0.5, 1.0), "Sent")
                    else:
                        imgui.text_colored((1.0, 0.2, 0.2, 1.0), "Failed")
                
                imgui.end_table()
            imgui.end_child()

        imgui.end_child()

        # Barra inferior (Subida a 60px para evitar cortes en Windows)
        imgui.set_cursor_pos((0, imgui.get_window_height() - 60))
        imgui.begin_child("BottomBar", (0, 60), False)
        imgui.separator()
        imgui.spacing()
        imgui.set_cursor_pos_x(15)
        imgui.text_colored((0.1, 0.8, 0.4, 1.0) if queue_manager.is_connected else (1, 0.2, 0.2, 1.0), 
                           f"{icons_fontawesome.ICON_FA_DATABASE}  REDIS STATUS")
        imgui.same_line(imgui.get_window_width() - 320)
        imgui.text_disabled(f"{icons_fontawesome.ICON_FA_MICROCHIP} CPU: {self.cpu_usage}%  |  {icons_fontawesome.ICON_FA_SERVER} RAM: {int(self.ram_usage)}MB")
        imgui.spacing(); imgui.spacing() # Espacio extra debajo de los indicadores
        imgui.end_child()

        # Modales
        if self.show_delete_confirm:
            imgui.open_popup("Confirm Delete")
            self.delete_target = self.show_delete_confirm
            self.show_delete_confirm = None
        if imgui.begin_popup_modal("Confirm Delete", True, imgui.WindowFlags_.always_auto_resize)[0]:
            imgui.text(f"Delete '{self.delete_target}'?")
            imgui.spacing()
            if imgui.button("CONFIRM", (100, 30)):
                config_manager.remove_client(self.delete_target)
                self.sessions = config_manager.get_client_list()
                imgui.close_current_popup()
            imgui.same_line()
            if imgui.button("CANCEL", (100, 30)):
                imgui.close_current_popup()
            imgui.end_popup()

        if self.show_add_modal:
            imgui.open_popup("Add New Client")
            self.show_add_modal = False
        if imgui.begin_popup_modal("Add New Client", True, imgui.WindowFlags_.always_auto_resize)[0]:
            imgui.text("Name:")
            _, self.new_client_name = imgui.input_text("##n", self.new_client_name)
            if imgui.button("Create"):
                if self.new_client_name:
                    config_manager.set_client_config(self.new_client_name, {"headless": True})
                    self.sessions = config_manager.get_client_list()
                imgui.close_current_popup()
            imgui.same_line()
            if imgui.button("Cancel"):
                imgui.close_current_popup()
            imgui.end_popup()

        if self.show_config_client:
            imgui.open_popup("Client Config")
            self.current_cfg_name = self.show_config_client
            self.headless_val = config_manager.get_client_config(self.current_cfg_name).get("headless", True)
            self.show_config_client = None
        if imgui.begin_popup_modal("Client Config", True, imgui.WindowFlags_.always_auto_resize)[0]:
            name = getattr(self, 'current_cfg_name', 'Unknown')
            imgui.text(f"Settings: {name}")
            _, self.headless_val = imgui.checkbox("Headless Mode", self.headless_val)
            if imgui.button("Save & Close"):
                config_manager.set_client_config(name, {"headless": self.headless_val})
                imgui.close_current_popup()
            imgui.end_popup()

        if self.show_ycloud_config:
            imgui.open_popup("Ycloud Configuration")
            self.show_ycloud_config = False
        
        # Le damos un tamaño inicial razonable al popup de Ycloud
        imgui.set_next_window_size((600, 500), imgui.Cond_.first_use_ever)
        if imgui.begin_popup_modal("Ycloud Configuration", True, imgui.WindowFlags_.no_saved_settings)[0]:
            imgui.text_colored((0.3, 0.7, 1.0, 1.0), "1. YCLOUD API CONFIGURATION")
            
            c1, self.ycloud_api_key = imgui.input_text("API Key", self.ycloud_api_key)
            c2, self.ycloud_url = imgui.input_text("API URL", self.ycloud_url)
            c3, self.ycloud_from = imgui.input_text("From Number", self.ycloud_from)
            c4, self.ycloud_cost_usd = imgui.input_float("Cost per Msg (USD)", self.ycloud_cost_usd, format="%.3f")
            c5, self.ycloud_exchange_pen = imgui.input_float("Exchange Rate (PEN)", self.ycloud_exchange_pen, format="%.2f")
            
            imgui.spacing(); imgui.separator(); imgui.spacing()
            imgui.text_colored((0.3, 0.7, 1.0, 1.0), "2. GLOBAL VARIABLES")
            c_col, self.ycloud_colegio = imgui.input_text("Colegio", self.ycloud_colegio)
            c_num, self.ycloud_numero = imgui.input_text("Número", self.ycloud_numero)
            
            imgui.spacing(); imgui.separator(); imgui.spacing()
            imgui.text_colored((0.3, 0.7, 1.0, 1.0), "3. MESSAGE TEMPLATE")
            imgui.text_disabled("Use variables: {Estudiante}, {Grado}, {Sección}, {Colegio}, {Numero}")
            
            c6, self.ycloud_template = imgui.input_text_multiline("##ycloudtpl", self.ycloud_template, (imgui.get_content_region_avail().x, 150))
            
            if c1 or c2 or c3 or c4 or c5 or c_col or c_num or c6:
                config_manager.settings["global"].update({
                    "ycloud_api_key": self.ycloud_api_key,
                    "ycloud_url": self.ycloud_url,
                    "ycloud_from": self.ycloud_from,
                    "ycloud_cost_usd": self.ycloud_cost_usd,
                    "ycloud_exchange_pen": self.ycloud_exchange_pen,
                    "ycloud_colegio": self.ycloud_colegio,
                    "ycloud_numero": self.ycloud_numero,
                    "ycloud_template": self.ycloud_template
                })
            
            imgui.spacing(); imgui.separator(); imgui.spacing()
            if imgui.button("SAVE & CLOSE", (150, 30)):
                config_manager.save()
                imgui.close_current_popup()
            imgui.same_line()
            if imgui.button("CANCEL", (100, 30)):
                imgui.close_current_popup()
                
            imgui.end_popup()

        imgui.end()
        imgui.pop_style_color(4)

def run_gui_app(loop):
    import imgui_bundle
    ui = WhatsAppUI(loop)
    params = hello_imgui.RunnerParams()
    params.app_window_params.window_title = "COLECHECK WSP-ADMIN"
    params.app_window_params.window_geometry.size = (1000, 650) # Un poco más grande para evitar scrolls
    params.callbacks.show_gui = ui.draw
    params.imgui_window_params.default_imgui_window_type = hello_imgui.DefaultImGuiWindowType.no_default_window
    
    # --- FIX ICON LOADING ---
    def load_fonts():
        try:
            font_path = "fonts/DroidSans.ttf"
            if os.path.exists(font_path):
                # Carga tu fuente local y automáticamente le adjunta los iconos FontAwesome
                hello_imgui.load_font_ttf_with_font_awesome_icons(font_path, 18.0)
            else:
                # Si no existe la fuente, usa la fuente predeterminada PERO con iconos activados
                hello_imgui.imgui_default_settings.load_default_font_with_font_awesome_icons()
        except Exception as e:
            print(f"⚠️ Error al cargar iconos: {e}")

    params.callbacks.load_additional_fonts = load_fonts
    immapp.run(params)
